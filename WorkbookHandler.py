from typing import Union
from datetime import datetime
from openpyxl import load_workbook, Workbook

class WorkbookHandler():
    def __init__(self):
        self._wb = None
        self._activeSheet = None
        self._activeSheetName = None
        self._activeSheetIndex = None

    def populateProjectNumber(self, project: Union[str, int]):
        """
        Populates project number
        :param project: Project number
        :return: None
        """
        self._activeSheet.cell(row=1, column=1).value = project

    def populateProjectName(self, projectName: str):
        """
        Populates project name
        :param projectName: Project name
        :return: None
        """
        self._activeSheet.cell(row=1, column=2).value = projectName

    def populateDate(self):
        """
        Populates date
        :return: None
        """
        self._activeSheet.cell(row=2, column=1).value = datetime.now().strftime("%B %d, %Y (%a)")

    def populateDailyINC(self, inc: float):
        """
        Populates Daily Incidence
        :param inc: Daily incedence
        :return: None
        """
        self._activeSheet.cell(row=2, column=18).value = inc

    def populateAVGDailyLOI(self, loi):
        """
        Populates Avg. Daily LOI
        :param loi: AVG Daily LOI
        :return: None
        """
        self._activeSheet.cell(row=3, column=18).value = loi

    def populateAVGOverallLOI(self, loi):
        """
        Populates Avg. Overall LOI
        :param loi: AVG Overall LOI
        :return: None
        """
        self._activeSheet.cell(row=4, column=18).value = loi

    def copyRows(self, count):
        """
        Extends rows
        :param count: Number of rows to copy
        :return: None
        """
        for i in range(count-1):
            row = i + 9
            self._activeSheet.cell(row=row, column=7).value = f"=(E{row} - F{row}) * 60"
            self._activeSheet.cell(row=row, column=9).value = f"=(H{row}) * 60"
            self._activeSheet.cell(row=row, column=10).value = f"=IFERROR(H{row}/E{row},0)"
            self._activeSheet.cell(row=row, column=11).value = f"=IFERROR(IF(G{row}>0,(G{row}+I{row})/(E{row}*60),I{row}/(E{row}*60)),0)"
            self._activeSheet.cell(row=row, column=13).value = f"=IFERROR(ROUND(E{row}*$U$1,2),0)"
            self._activeSheet.cell(row=row, column=14).value = f"=L{row}-M{row}"
            self._activeSheet.cell(row=row, column=15).value = f"=RANK(N{row},$N$8:$N$140)"
            self._activeSheet.cell(row=row, column=16).value = f"=IFERROR(T{row}/$AC$85,0)"
            self._activeSheet.cell(row=row, column=18).value = f"=IFERROR(L{row}/E{row},0)"
            self._activeSheet.cell(row=row, column=22).value = f"=IFERROR(T{row}/E{row},0)"

    def setWorkbook(self, path):
        """
        Sets workbook
        :param path: Filepath
        :return: None
        """
        self._wb = load_workbook(path)

    def setActiveSheet(self, activeSheet: str) -> None:
        """
        Sets active sheet
        :param active: Sheet name
        :return: None
        """
        self._activeSheet = self._wb[activeSheet]

    def setActiveSheetName(self, sheetName: str) -> str:
        """
        Sets active sheet name
        :param sheetName: Sheet name
        :return: String
        """
        self._activeSheetName = sheetName
        return sheetName

    def getWorkbook(self):
        if self._wb is None:
            return 'No active workbook'
        return self._wb

    def getActiveSheet(self):
        if self._activeSheet is None:
            return 'No active sheet'
        return self._activeSheet

    def getActiveSheetName(self):
        if self._activeSheetName is None:
            return 'No active sheet name'
        return self._activeSheetName

    def save(self, projectid):
        # TODO Replace with path instead of title
        self._wb.save(f"{projectid}_Production_ReportTEST.xlsx")
