import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showerror
import os
import shutil
from openpyxl import load_workbook
from datetime import datetime
import SQLDictionary
import DataPuller
import WorkbookHandler
from configparser import ConfigParser


config_object = ConfigParser()
config_object.read("config.ini")
qry_info = config_object["SQL CODE"]
file_paths = config_object['FILE PATHS']


sqld = SQLDictionary.SQLDictionary()
dpull = DataPuller.DataPuller()
wh = WorkbookHandler.WorkbookHandler()


def create_path(projectid):
    if os.path.exists(f"{file_paths['SRC']}{projectid}/PRODUCTION/") == False:
        src = f"{file_paths['SRC']}PRODUCTION/BLANK Production.xlsx"
        os.mkdir(f"{file_paths['SRC']}{projectid}/PRODUCTION/")
        dst = f"{file_paths['SRC']}{projectid}/PRODUCTION/{projectid}_Production_ReportTEST.xlsx"
        shutil.copy(src, dst)
        del src, dst
    elif os.path.exists(f"{file_paths['SRC']}{projectid}/PRODUCTION/{projectid}_Production_ReportTEST.xlsx") == False:
        src = f"{file_paths['SRC']}PRODUCTION/BLANK Production.xlsx"
        dst = f"{file_paths['SRC']}{projectid}/PRODUCTION/{projectid}_Production_ReportTEST.xlsx"
        shutil.copy(src, dst)
        del src, dst
    else:
        pass

    return f"{file_paths['SRC']}{projectid}/PRODUCTION/{projectid}_Production_ReportTEST.xlsx"


def read_excel(path: str, sheetname, projectid):
    wh.setWorkbook(path)
    wh.setActiveSheet(wh.getWorkbook().sheetnames[0])
    print(wh.getActiveSheet().title)
    wh.copyRows(10)
    wh.save(projectid)
    # wb = load_workbook(path)
    # sheet1 = wb[wb.sheetnames[0]]
    # print(sheetname)
    # print()
    # print(sheet1.title)
    # print()



# create_path(12523)
active_id_df = dpull.pull_data(sqld.project_id_list(qry=qry_info["ActiveProjectIDs"]))
active_dict = dict.fromkeys(active_id_df['projectid'])

active_dict = {'12523': '12523',
               '12523C': '12523'}

prev = None
for key in active_dict:
    active_dict[key] = key[:5]
    if prev is None or prev != active_dict[key]:
        path = create_path(active_dict[key])

    elif prev == active_dict[key]:
        print(f'RUNNING {active_dict[key]} again')
    prev = active_dict[key]
    read_excel(path, key, active_dict[key])
    # print(path)




# keys = list(active_dict.keys())
# prev = None
# for key in active_dict:
#     if prev is not None or prev != key:
#         create_path(active_dict[key])
#     prev = key

print(active_dict)
active_id_list_details = list(active_id_df['projectid'])
active_id_list = [projectid[:5] for projectid in active_id_list_details]


if __name__ == '__main__':
    print()

